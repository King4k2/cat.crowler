import time
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import json
import shutil
import os


def gen_url1(code, raw_url="https://parts.cat.com/en/catcorp/", mode="1"):
    code = str(code)
    if len(code) == 6:
        code = code[:2] + "-" + code[2:]
        url = raw_url + code
        if mode == "2":
            return code
        return url
    elif len(code) == 7:
        code = code[:3] + "-" + code[3:]
        url = raw_url + code
        if mode == "2":
            return code
        return url
    else:
        print(code)
        return None


def main():
    workbook = load_workbook("Full list.xlsx")
    sh = workbook["Sheet1"]
    k_ = 0
    for a in range(1+1, sh.max_row):
        print(f"Progress: {a}/{sh.max_row}")
        if sh[f"D{a}"].value != "":
            continue
        item_id = sh[f"A{a}"].value
        part_id = sh[f"B{a}"].value
        try:
            part_id = int(part_id)
        except Exception as err:
            pass
        part_id = str(part_id).replace(" ", "")
        print(part_id)
        if 5 < len(str(part_id)) < 8:
            pass
        else:
            if "/" not in str(part_id):
                sh[f"D{a}"] = "Not CAT parts"
                continue
            else:
                sh[f"B{sh.max_row+1}"] = str(part_id).split("/")[1]
                sh[f"B{a}"] = str(part_id).split("/")[0]
                part_id = str(part_id).split("/")[0]
        # --------------------------------------------------------------
        proxies = {
            "https": "",
            # "http": ""
        }
        headers = {
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8"
                      ",application/signed-exchange;v=b3;q=0.7",
            "Accept-Encoding": "gzip, deflate, br, zstd",
            "Accept-Language": "en-US;q=0.8,en;q=0.7",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/124.0.0.0 YaBrowser/24.6.0.0 Safari/537.36",
            "Sec-Ch-Ua-Mobile": "?0",
            "Sec-Ch-Ua-Platform": '"Windows"',
            "Sec-Fetch-Dest": "document",
            "Sec-Ch-Ua": '"Chromium";v="124", "YaBrowser";v="24.6", "Not-A.Brand";v="99", "Yowser";v="2.5"'
        }
        # --------------------------------------------------------------

        try:
            url = gen_url1(code=part_id)
            if url is None:
                continue
            l_ = 0
            while True:
                req = requests.get(url, headers=headers)
                try:
                    if int(req.status_code) == 200:
                        break
                    if int(req.status_code) == 404:
                        if l_ > 1:
                            sh[f"D{a}"] = "Not CAT parts"
                            break
                        else:
                            l_ = l_ + 1
                    else:
                        time.sleep(1)
                        continue
                except Exception as err:
                    if str(err).startswith("SOCKSHTTPSConnectionPool"):
                        continue
                    else:
                        pass
            if l_ > 1:
                continue
            soup = BeautifulSoup(req.text, "lxml")
            part_desc = soup.find("title").text.split(" | ")[0]
            page_id = soup.find("meta", {"name": "pageId"}).get("content")
            sh[f"D{a}"] = part_desc
            url = (f"https://parts.cat.com/api/product/detail?productId={page_id}&storeIdentifier=CATCorp&locale=en_US"
                   f"&partNumber={gen_url1(code=part_id, mode="2")}&&storeId=21801&langId=-1")
            while True:
                try:
                    req = requests.get(url=url, headers=headers)
                    if int(req.status_code) == 200:
                        break
                    else:
                        time.sleep(1)
                        continue
                except Exception as err:
                    if str(err).startswith("SOCKSHTTPSConnectionPool"):
                        print(err)
                        continue
                    else:
                        print(err)
            resp_ = json.loads(req.content)
            if str(req.content).find("compatible") == -1:
                compatible_string = "No Data"
            else:
                compatible_string = ""
                for compatible_category in resp_["compatibleModels"]:
                    compatible_string = compatible_string + compatible_category + ": "
                    for compatible in resp_["compatibleModels"][compatible_category]:
                        compatible_string = compatible_string + compatible
                        if resp_["compatibleModels"][compatible_category][-1] != compatible:
                            compatible_string = compatible_string + ", "
                    compatible_string = compatible_string + "; " + "\n"

            sh[f"F{a}"] = compatible_string
        except Exception as err:
            pass
        finally:
            if a//100 > 0+k_:
                print(f"Copy was saved to 'cat.crowler/copies/Full list_copy{a}.xlsx'")
                workbook.save("Full list.xlsx")
                shutil.copy("Full list.xlsx", f"copies/Full list_copy{a}.xlsx")
                k_ = k_ + 1
    workbook.close()
    print("Finished!")


if __name__ == "__main__":
    main()
