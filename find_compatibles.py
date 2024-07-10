import openpyxl


def main():
    workbook_of_compatibles = openpyxl.load_workbook("Book4.xlsx")
    wb = openpyxl.load_workbook("Full list.xlsx")
    sheet_of_compatibles = workbook_of_compatibles["Sheet1"]
    sh = wb["Sheet1"]
    p_ = 0
    for i in range(2, sh.max_row):
        print(f"process: {i}/{sh.max_row}")
        general_compatibles_check_list = []
        general_compatibles = ""
        written_compatibles = str(sh[f"F{i}"].value).replace(" ", "")
        written_compatibles_in_rows_list = written_compatibles.split("\n")
        for written_compatibles_in_row in written_compatibles_in_rows_list:
            if written_compatibles_in_row == "\n" or written_compatibles_in_row == "" or written_compatibles_in_row == 'None':
                continue
            try:
                written_compatibles_list_raw = written_compatibles_in_row.split(":")[1].split(",")
            except IndexError:
                continue
            for written_compatible in written_compatibles_list_raw:
                for b in range(1, sheet_of_compatibles.max_row):
                    compatible = str(sheet_of_compatibles[f"A{b}"].value).replace(" ", "")
                    if written_compatible.replace(" ", "").replace(";", "") == compatible:
                        general_compatibles_check_list.append(compatible)
                        break
                continue
        for m in general_compatibles_check_list:
            check_flag = 0
            try:
                for n in general_compatibles.split(", "):
                    if m == n:
                        check_flag = check_flag + 1
            except Exception as err:
                print(err)
            if check_flag < 1:
                general_compatibles = general_compatibles + m + ", "
        sh[f"E{i}"] = general_compatibles
        if (i//500)-p_ > 0:
            p_ = p_ + 1
            wb.save("Full list.xlsx")
    wb.close()
    workbook_of_compatibles.close()


if __name__ == "__main__":
    main()
